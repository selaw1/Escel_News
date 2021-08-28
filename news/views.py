from django.shortcuts import render
from django.conf import settings
from django.templatetags.static import static
from django.contrib.staticfiles.storage import staticfiles_storage
from django.contrib.staticfiles.finders import find

import openpyxl


def home_news(request):
    return render(request, 'news/home.html')    

def get_static(path):
<<<<<<< HEAD
    if settings.DEBUG:
        return find(path)
    else:
        # print('static: ', static(path))
        # print('staticfilestorage: ', staticfiles_storage.path(path))
        return staticfiles_storage.path(path)
=======
#     if settings.DEBUG:
#         return find(path)
#     else:
    print('static: ', static(path))
    print('staticfilestorage: ', staticfiles_storage.path(path))
    return staticfiles_storage.path(path)
>>>>>>> origin/main

def excel(path):

    wb = openpyxl.load_workbook(path)

    # getting a particular sheet by name out of many sheets
    # worksheet = wb["Sheet2"]
    worksheet = wb.active
    max_row = worksheet.max_row

    row_values = []
    row_numbers = []

    count = 0
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=1):
        for cell in row:
            count += 1

            row_values.append(cell.value)
            row_numbers.append(count)

    excel_data = dict(zip(row_numbers,row_values))
    return excel_data

def sport_news(request):
    path = 'excel/sport_update.xlsx'
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Sport News'})

def business_news(request):
    path = "excel/business.xlsx"
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Business News'})

def health_news(request):
    path = 'excel/health.xlsx'
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Health News'})

def tech_news(request):
    path = 'excel/technology.xlsx'
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Technology News'})

def politics_news(request):
    path = 'excel/politics.xlsx'
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Politic News'})

def culture_news(request):
    path = 'excel/culture.xlsx'
    excel_data = excel(get_static(path))
    return render(request, 'news/news.html', {"excel_data":excel_data, 'type':'Culture News'})
